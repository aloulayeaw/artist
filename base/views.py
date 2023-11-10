from django.shortcuts import render
from .models import *
#from dashboard.models import Blog
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
# Create your views here.
from openpyxl import load_workbook
from django.shortcuts import render
#from .models import Song
from django.conf import settings
from paydunya import InvoiceItem, Store
from paydunya import Store
from django.shortcuts import redirect
from django.http import HttpResponse
#from .forms import PaymentForm
import paydunya
from django.conf import settings
from django.shortcuts import render, redirect
#from .forms import PaymentForm
import stripe
stripe.api_key = "sk_test_51LuypqEMbpaxmGP6WCG43ONNmFMRfyKuOxPihh9OU3UJVYc72zAyV0oU7KmQCcjclpdNemi6kbP9c7aNyeWgW5Hh00jCTh8xsV"
from django.core.mail import send_mail
from django.http import JsonResponse
import requests
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
#from .models import PaymentForm
#from .forms import PaymentFormModelForm
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
import json
from .models import OrderCounterModel

paydunya.debug = False

paydunya.api_keys = settings.PAYDUNYA_ACCESS_TOKENS

store = Store(name='Keulthieu')



def home(request):
    artist_name = "Keulthieu The Name"

    # Effectuez une requête à l'API iTunes Search avec une limite de 5 résultats
    url = f"https://itunes.apple.com/search?term=Keulthieu The Name&entity=song&limit=7"
    response = requests.get(url)

    # Vérifiez si la requête a réussi
    if response.status_code == 200:
        # Analysez la réponse JSON
        data = response.json()
        songs = data.get("results", [])

        # Triez les chansons en filtrant celles ayant une date de sortie valide
        songs = [song for song in songs if song.get("releaseDate")]

        # Triez les chansons par releaseDate du plus récent au plus ancien
        songs = sorted(songs, key=lambda x: x.get("releaseDate"), reverse=True)

        # Sélectionnez les 5 premières chansons
        songs = songs[:7]

        # Envoyez les données triées à votre modèle HTML
        context = {
            "artist_name": artist_name,
            "songs": songs,
        }

        return render(request, 'index.html', context)
    else:
        # Gérez les erreurs de requête ici
        error_message = "Une erreur s'est produite lors de la récupération des données iTunes."
        return render(request, 'error.html', {'error_message': error_message})
  
def get_itunes_data(request):
    artist_name = "Keulthieu The Name"
    url = f"https://itunes.apple.com/search?term={artist_name}&entity=album"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        albums = data.get('results', [])
        return JsonResponse({'albums': albums})
    else:
        return JsonResponse({'error': 'Unable to fetch data'})
    
def paiement(request):
    
    return render(request, 'base/paiement.html')

def payment_view(request):
#     if request.method == 'POST':
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             # Récupérez les données du formulaire
#             first_name = form.cleaned_data['first_name']
#             last_name = form.cleaned_data['last_name']
#             email = form.cleaned_data['email']
#             phone = form.cleaned_data['phone']
            
#             # Créez une charge avec Stripe (c'est un exemple simple, veuillez personnaliser selon vos besoins)
#             charge = stripe.Charge.create(
#                 amount=1000,  # Montant en cents (10 $)
#                 currency='usd',
#                 description='Paiement pour votre commande',
#                 source=request.POST['stripeToken'],  # Stripe token
#             )
            
#             # Si le paiement est réussi, redirigez l'utilisateur vers /home avec un message de succès
#             if charge.status == 'succeeded':
#                 return redirect('/home/?success=True')

#     else:
#         form = PaymentForm()

      return render(request, 'base/paiement.html')

def effectuer_transaction(request):
    # Définissez les clés d'accès Paydunya
    items = [
        InvoiceItem(
            name='Album Keulthieu',
            quantity=1,
            unit_price="10000",  # Mettez le prix unitaire en tant que chaîne
            total_price="10000",  # Le total est le produit du prix unitaire et de la quantité
            description='Test',
    )]
    
    # Créez la facture
    invoice = paydunya.Invoice(store)
    invoice.add_items(items)

    # Effectuez la transaction
    successful, response = invoice.create()
    
    if successful:
        # La transaction a réussi, vous pouvez rediriger l'utilisateur ou afficher un message de succès
        return HttpResponse('<h2>Merci pour le paiement</h2>')
    else:
        # Gérez les erreurs de la transaction (par exemple, affichez un message d'erreur)
        return HttpResponse('<h2>La transaction a échoué</h2>')

# def payment_form_view(request):
#     if request.method == 'POST':
#         form = PaymentFormModelForm(request.POST)
#         if form.is_valid():
#             form.save()  # Sauvegarde du formulaire dans la base de données
#             return redirect('/')
#     else:
#         form = PaymentFormModelForm()

#     return render(request, 'base/paiement.html', {'form': form})
def payment_form_view(request):
    if request.method == 'POST':
        name = request.POST.get('Name', '')
        email = request.POST.get('Email', '')
        phone = request.POST.get('phone', '')
        message = request.POST.get('adresse', '')
        product = request.POST.get('product', '')

        if not name or not email or not phone or not message:
            return JsonResponse({'error_message': 'Tous les champs sont obligatoires'})

        email_body = f"Nom: {name}\nEmail: {email}\nTéléphone: {phone}\nProduct: {product}\nMessage: {message}"



        send_mail(
            'Commande Mixtape Demb Ak Tay',
            email_body,
            'zblackofficiel@gmail.com',
            ['zblackofficiel@gmail.com'],
            fail_silently=False,
            #
        )

        data = {
            "item_name": "Mistake Demb ak Tey",
            "item_price": "10000",
            "currency": "XOF",
            "ref_command": email+phone, 
            "command_name": "Mistake Demb ak Tey",
            "ipn_url": "https://www.keulthieuthename.com/",
            "success_url": "https://www.keulthieuthename.com/",
            "cancel_url": "https://www.keulthieuthename.com/base",
        }
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded;charset=utf-8",
            "API_KEY": "97fdd2a3880062cd590338038fca0ffbf4199e5bf2ff6405d28dc8ded8b2bd82",
            "API_SECRET": "cf4389bf11b703fb67b76b1c8ec68de4d5f909fa3a48449a01097a27ad904eda",
        }


        response = requests.post(
            url="https://paytech.sn/api/payment/request-payment",
            data=data,
            headers=headers,
        )
        
        #print("response ", response.text)

        if response.status_code == 200:
            redirect_url = response.json().get("redirect_url")
            if redirect_url:
                return HttpResponseRedirect(redirect_url)  # Redirection vers le redirect_url
            else:
                return JsonResponse({"success": False, "error_message": "Le 'redirect_url' n'a pas été trouvé dans la réponse JSON"})
        else:
            return JsonResponse({"success": False, "status_code": response.status_code, "error_message": response.text})

def contact(request):
    if request.method == 'POST':
        name = request.POST.get('Name')
        email = request.POST.get('Email')
        phone = request.POST.get('phone')
        message = request.POST.get('message')

        # Validation du formulaire (vous pouvez ajouter plus de validation au besoin)
        if name and email and phone and message:
            # Envoyer l'e-mail
            subject = 'Nouveau message de contact'
            message_text = f'Nom: {name}\nEmail: {email}\nTéléphone: {phone}\nMessage: {message}'
            from_email = 'zblackofficiel@gmail.com'  # Remplacez par votre propre adresse e-mail
            recipient_list = ['zblackofficiel@gmail.com']  # Adresse e-mail de destination

            try:
                send_mail(subject, message_text, from_email, recipient_list, fail_silently=False)
                response_data = {'success': True, 'message': 'Les données ont été envoyées avec succès.'}
                return JsonResponse(response_data)
            except Exception as e:
                response_data = {'success': False, 'message': 'Une erreur s\'est produite lors de l\'envoi de votre message.'}
                return JsonResponse(response_data, status=500)
        else:
            response_data = {'success': False, 'message': 'Veuillez remplir tous les champs du formulaire.'}
            return JsonResponse(response_data, status=400)

    # Si la méthode n'est pas POST, affichez simplement le formulaire
    return render(request, 'index.html')#return render(request, 'payment_form.html')


#     filepath = r"D:\Dev\Website Communautaire\webCommune\webCommune\static\assets\fichiers\liste_des_premiers_compagnons_de_SLLASW.xlsx"
#     workbook = load_workbook(filepath, read_only=True)
#     sheet = workbook.active

#     data = []
#     data2 = {}
#     data3 = {} 
#     data4 = {} 

#     for row in sheet.iter_rows(values_only=True):
#         if len(row) >= 12 and row[0] != "prenom" and row[6] and row[7] and row[8] and row[9] and row[10] and row[11]:
#             data.append({
#                 "prenom": row[0],
#                 "nom": row[1],
#                 "latitude": row[6],
#                 "longitude": row[7]
#             })

#             if row[3] in data2:
#                 data2[row[3]]["count"] += 1
#             else:
#                 data2[row[3]] = {
#                     "count": 1,
#                     "latitudeDep": row[8],
#                     "longitudeDep": row[9]
#                 }
#             if row[4] in data3:
#                 data3[row[4]]["number"] += 1
#             else:
#                 data3[row[4]] = {
#                     "number": 1,
#                     "latituderReg": row[10],
#                     "longitudeReg": row[11]
#                 }
                
#             if row[4] in data4:
#                 data4[row[4]].append(
#                     {"nom": row[1], 
#                     "prenom": row[0],
#                     "latitudeReg": row[10],
#                     "longitudeReg": row[11]
#                     }
#                 )
#             else:
#                 data4[row[4]] = [
#                     {"nom": row[1], 
#                     "prenom": row[0],
#                     "latitudeReg": row[10],
#                     "longitudeReg": row[11]
#                     }
#                 ]

            
#             #print("data2",data4)

#     return render(request, 'base/cartographie.html', {"data": data, "data2": data2,"data3": data3,"data4": data4})

# def fetch_songs(request):
#     artist_name = "keulthieu the name"  # Remplacez par le nom de l'artiste que vous souhaitez rechercher

#     # Recherchez les chansons de l'artiste sur iTunes
#     itunes = pyitunes.PyiTunes()
#     songs = itunes.search_artist(artist_name)

#     # Enregistrez les chansons dans la base de données
#     for song_info in songs:
#         song = Song(
#             title=song_info.get("trackName", ""),
#             artist=song_info.get("artistName", ""),
#             album=song_info.get("collectionName", ""),
#             url=song_info.get("previewUrl", ""),
#         )
#         song.save()

#     # Récupérez toutes les chansons de l'artiste depuis la base de données
#     songs = Song.objects.filter(artist=artist_name)

#     return render(request, "base/base.html", {"songs": songs})